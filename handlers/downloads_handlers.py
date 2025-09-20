from peewee import DoesNotExist
from datetime import datetime, timedelta
import openpyxl
from io import BytesIO
from telebot import TeleBot
from telebot.types import Message, ReplyKeyboardRemove
from config import admin_id
from DATABASE.peewee_config import Client, Week, Lesson
from states import reg_states_admin
from keyboards.main_keyboards import (
    go_to_menu,
    main_admin_commands,
    downloads_type)

def reg_downloads_handlers(bot: TeleBot):
    """
    Функция для регистрации обработчиков выгрузки данных
    :param bot: переменная с приложением
    :return: None
    """

    @bot.message_handler(state=reg_states_admin.in_any_block,
                         func=lambda message: 'Выгрузка данных' in message.text)
    def downloads_actions(message: Message):
        """
        Обработчик состояния in_any_block при нажатии на кнопку "Выгрузка данных".
        Предлагает админу выбрать тип выгрузки данных, воспользовавшись кнопками.
        Меняет состояние на reg_states_admin.in_downloads
        :param message: переменная с приложением
        :return: None
        """

        bot.send_message(message.chat.id,
                         'Выберите тип выгрузки',
                         reply_markup=downloads_type(), parse_mode='HTML')
        bot.set_state(message.from_user.id, reg_states_admin.in_downloads, message.chat.id)

    @bot.message_handler(state=[reg_states_admin.in_downloads,
                                reg_states_admin.in_downloads_schedule],
                         func=lambda message: 'Перейти в основное меню' in message.text)
    def return_to_menu(message: Message):
        """
        Обработчик нажатия кнопки "Перейти в основное меню" в состояниях in_downloads, in_downloads_schedule.
        Возвращает пользователя в основное меню.
        Меняет состояние на reg_states_admin.in_any_block
        :param message: переменная с приложением
        :return: None
        """

        bot.send_message(message.chat.id,
                         'Выберите действие',
                         reply_markup=main_admin_commands(), parse_mode='HTML')
        bot.set_state(message.from_user.id, reg_states_admin.in_any_block, message.chat.id)

    @bot.message_handler(state=reg_states_admin.in_downloads,
                         func=lambda message: 'Данные активных клиентов' in message.text)
    def downloads_clients(message: Message):
        """
        Обработчик состояния после нажатия на кнопку "Данные активных клиентов" в меню выгрузки.
        Создает Exsel файл в памяти, создает лист clients_sheet,
        Заполняет лист данными клиентов.
        Сохраняет измененный Exsel файл в памяти и отправляет его админу в чат.
        Обрабатывает ошибки.
        Меняет состояние пользователя на reg_states_admin.admin_menu
        :param message: переменная с приложением
        :return: None
        """

        try:
            clients_data = Client.select()
            weeks = Week.select()

            # Создаем Excel файл в памяти
            workbook = openpyxl.Workbook()
            clients_sheet = workbook.active
            clients_sheet.title = 'Дынные клиентов'

            # Заполняем лист с данными активных клиентов
            headers = ['№ п/п', 'ID клиента', 'Имя', 'Фамилия', 'Номер телефона', 'Имя ребенка', 'Дата рождения ребенка']
            for col_num, header in enumerate(headers, 1):
                clients_sheet.cell(row=1, column=col_num, value=header)

            for i_index, i_client in enumerate(clients_data, 2):
                clients_sheet.cell(row=i_index, column=1, value=i_index - 1)
                clients_sheet.cell(row=i_index, column=2, value=i_client.clients_id)
                clients_sheet.cell(row=i_index, column=3, value=i_client.clients_name)
                clients_sheet.cell(row=i_index, column=4, value=i_client.clients_sirname)
                clients_sheet.cell(row=i_index, column=5, value=i_client.clients_number)
                clients_sheet.cell(row=i_index, column=6, value=i_client.clients_child_name)
                clients_sheet.cell(row=i_index, column=7, value=i_client.clients_child_birthday)

            excel_file = BytesIO()
            workbook.save(excel_file)
            excel_file.seek(0)

            bot.send_document(message.chat.id,
                              document=excel_file,
                              visible_file_name='output_data_clients.xlsx',
                              caption='В файле "output_data_clients.xlsx" выгрузка из базы данных по активным клиентам',
                              reply_markup=go_to_menu(), parse_mode='HTML')

            bot.set_state(message.from_user.id, reg_states_admin.admin_menu, message.chat.id)
        except Exception as e:
            bot.send_message(message.chat.id, f"❌ Ошибка при выгрузке: {str(e)}\n"
                                              f"Для возврата в меню воспользуйтесь кнопкой ниже", reply_markup=go_to_menu(), parse_mode='HTML')
            bot.set_state(message.from_user.id, reg_states_admin.admin_menu, message.chat.id)

    @bot.message_handler(state=reg_states_admin.in_downloads,
                         func=lambda message: 'Расписания' in message.text)
    def downloads_actions(message: Message):
        """
        Обработчик состояния после нажатия кнопки "Расписания" в меню выгрузки.
        Предлагает админу ввести дату понедельника недели, выгрузку расписания по которой он хочет получить.
        Меняет состояние на reg_states_admin.in_downloads_schedule
        :param message: переменная с приложением
        :return: None
        """

        bot.send_message(message.chat.id,
                         'Пришлите дату понедельника недели для выгрузки. Формат - DD.MM.YYYY',
                         reply_markup=ReplyKeyboardRemove(), parse_mode='HTML')
        bot.set_state(message.from_user.id, reg_states_admin.in_downloads_schedule, message.chat.id)

    @bot.message_handler(state=reg_states_admin.in_downloads_schedule)
    def downloads_clients(message: Message):
        """
        Создает Exsel файл в памяти, заполняет его лист schedule_sheet данными с уроками данной недели,
        сохраняет лист и файл в памяти, отправляет exsel файл админу.
        Обрабатывает ошибкаи.
        Меняет состояние на reg_states_admin.admin_menu
        :param message: переменная с приложением
        :return: None
        """

        try:
            #НЕОБХОДИМО ДОБАВИТЬ ПРОВЕРКУ ФОРМАТА ДАТЫ
            input_date = datetime.strptime(message.text, '%d.%m.%Y').date()
            week = Week.get_or_none(Week.monday_date == input_date)
            if week is None:
                raise DoesNotExist('Такой недели нет в базе данных.')
            next_day = week.monday_date
            lessons = week.lessons

            # Создаем Excel файл в памяти
            workbook = openpyxl.Workbook()
            schedule_sheet = workbook.active
            schedule_sheet.title = f'Расписание {week.monday_date}'

            # Заполняем лист с данными запрошенного расписания
            headers = ['День недели',
                       'Дата',
                       '8:00',
                       '8:45',
                       '9:30',
                       '10:15',
                       '11:00',
                       '11:45',
                       '12:30',
                       '13:15',
                       '17:30',
                       '18:15',
                       '19:00'
                       ]
            for col_num, header in enumerate(headers, 1):
                schedule_sheet.cell(row=1, column=col_num, value=header)

            for i_row in range(2, 8):
                schedule_sheet.cell(row=i_row, column=1, value=Lesson.days_dict[i_row - 2])
                schedule_sheet.cell(row=i_row, column=2, value=next_day)
                for i_lesson in lessons:
                    if i_lesson.day_of_week == i_row - 2:
                        client_data = [i_lesson.client.clients_name, i_lesson.client.clients_sirname]
                        schedule_sheet.cell(row=i_row, column=i_lesson.lesson_number + 2, value=' '.join(client_data))
                next_day += timedelta(days=1)

            schedule_sheet.cell(row=8, column=1, value='Дата понедельника: ')
            schedule_sheet.cell(row=8, column=2, value=input_date)

            excel_file = BytesIO()
            workbook.save(excel_file)
            excel_file.seek(0)

            bot.send_document(message.chat.id,
                              document=excel_file,
                              visible_file_name='output_data_schedule.xlsx',
                              caption=f'В файле "output_data_schedule.xlsx" вся выгрузка из базы данных за выбранную ({message.text}) неделю',
                              reply_markup=go_to_menu(), parse_mode='HTML')

            bot.set_state(message.from_user.id, reg_states_admin.admin_menu, message.chat.id)
        except Exception as e:
            bot.send_message(message.chat.id, f"❌ Ошибка при выгрузке: {str(e)}\nВведите корректную дату. Формат - DD.MM.YYYY\n"
                                              f"\n\nДля возврата в меню воспользуйтесь кнопкой ниже",
                             reply_markup=go_to_menu(), parse_mode='HTML')

    @bot.message_handler(state=reg_states_admin.in_downloads,
                         func=lambda message: 'Архивные данные' in message.text)
    def downloads_actions(message: Message):
        """
        Обработчик состояния после нажатия кнопки "Архивные данные" в меню выгрузки.
        Раздел в разработке, в дальнейшем тут будет раздел просмотра архивных данных.
        :param message: переменная с приложением
        :return: None
        """

        bot.send_message(message.chat.id,
                         'IN CONSTRACTION! PLS, RETURN TO MENU',
                         reply_markup=go_to_menu(), parse_mode='HTML')
        bot.set_state(message.from_user.id, reg_states_admin.admin_menu, message.chat.id)