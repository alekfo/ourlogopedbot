from datetime import datetime
import openpyxl
from io import BytesIO
from telebot import TeleBot
from telebot.types import Message
from DATABASE.peewee_config import Client, Week, Lesson
from states import reg_states_admin
from keyboards.main_keyboards import (
    go_to_menu,
    main_admin_commands,
    schedule_menu)


def reg_schedule_handlers(bot: TeleBot):
    """
    Функция для регистрации обработчиков для управления расписанием
    :param bot: переменная с приложением
    :return: None
    """

    @bot.message_handler(state=reg_states_admin.in_any_block,
                         func=lambda message: 'Управление расписанием' in message.text)
    def schedule_actions(message: Message):
        """
        Обработчик состояния reg_states_admin.in_any_block при нажатии на кнопку "Управление расписанием".
        Предлагает админу выбрать действие на расписанием.
        Меняет состояние пользователя на reg_states_admin.in_schedule
        :param message: сообщения от пользователя
        :return: None
        """

        bot.send_message(message.chat.id,
                         'Выберите действие с расписанием',
                         reply_markup=schedule_menu(), parse_mode='HTML')
        bot.set_state(message.from_user.id, reg_states_admin.in_schedule, message.chat.id)

    @bot.message_handler(state= [reg_states_admin.in_schedule,
                                 reg_states_admin.process_file,
                                 reg_states_admin.show_schedule,
                                 reg_states_admin.delete_client,
                                 reg_states_admin.mass_mailing_state],
                        func=lambda message: 'Перейти в основное меню' in message.text)
    def return_to_menu(message: Message):
        """
        Обработчик сообщения для перехода в основное меню из состояний:
                                reg_states_admin.in_schedule,
                                 reg_states_admin.process_file,
                                 reg_states_admin.show_schedule,
                                 reg_states_admin.delete_client,
                                 reg_states_admin.mass_mailing_state.
        Аналог обработчика admin_menu.
        Меняет состояние пользователя на reg_states_admin.in_any_block
        :param message: сообщения от пользователя
        :return: None
        """

        bot.send_message(message.chat.id,
                         'Выберите действие',
                         reply_markup=main_admin_commands(), parse_mode='HTML')
        bot.set_state(message.from_user.id, reg_states_admin.in_any_block, message.chat.id)

    @bot.message_handler(state=reg_states_admin.in_schedule,
                         func=lambda message: 'Добавить расписание' in message.text)
    def add_schedule(message: Message):
        """
        Обработчки состояния reg_states_admin.in_schedule при нажатии админа на кнопку "Добавить расписание".
        Предлагает админу прислать файл формата .xlsx для массовой загрузки недельного расписания.
        Меняет состояние пользователя на reg_states_admin.process_file
        :param message: сообщения от пользователя
        :return: None
        """

        bot.send_message(message.chat.id,
                         'Отправьте файл формата <b>.xlsx</b>\n'
                         'В строках файла должны содержаться дни недели (кроме воскресенья),\n'
                         'в столбцах - разбивка по урокам\n\n'
                         'Для избежания ошибок при загрузке рекомендуется ознакомиться с образцом загрузочного файла',
                         reply_markup=go_to_menu(), parse_mode='HTML')
        bot.set_state(message.from_user.id, reg_states_admin.process_file, message.chat.id)

    @bot.message_handler(content_types=['document'], state=reg_states_admin.process_file)
    def file_procced(message: Message):
        """
        Обработчик состояния reg_states_admin.process_file при получении файла от пользователя.
        Загружает Exsel файл для обработки.
        Добавляет в базу данных недельное расписание из файла: создает модели недели
        и каждого активного урока этой недели в peewee БД.
        Обрабатывает ошибки.
        В случае успешного добавления в БД выдает админу подтверждение и выводит основное расписание на экран.
        Меняет состояние пользователя на reg_states_admin.admin_menu
        :param message: сообщения от пользователя
        :return: None
        """

        try:
            # Получаем информацию о файле
            file_info = bot.get_file(message.document.file_id)
            downloaded_file = bot.download_file(file_info.file_path)
            # Создаем BytesIO объект для работы с файлом в памяти
            excel_data = BytesIO(downloaded_file)

            # Загружаем Excel файл в
            workbook = openpyxl.load_workbook(excel_data)
            sheet = workbook.active
            monday_data = sheet.cell(2, 2).value                        #забираем значение из ячейки
            # formated_data = datetime.strptime(monday_data, "%d.%m.%Y").date()      #преобразуем в DateField

            existing_week = Week.get_or_none(Week.monday_date == monday_data)
            if existing_week:
                # Удаляем все уроки этой недели
                Lesson.delete().where(Lesson.weekly_schedule == existing_week).execute()
                existing_week.delete_instance()                 # Удаляем для избежания повтора уникальности

            curr_week = Week.create(monday_date=monday_data)
            for i_row in range(2, 8):
                for i_col in range(3, 14):
                    cell = sheet.cell(i_row, i_col).value
                    if cell is None:
                        continue
                    else:
                        cell_list = cell.split()
                        name, sirname = cell_list
                        curr_client = Client.get_or_none((Client.clients_name == name) & (Client.clients_sirname == sirname))
                        if curr_client:
                            Lesson.create(
                                lesson_date=sheet.cell(row=i_row, column=2).value,
                                weekly_schedule=curr_week,
                                client=curr_client,
                                day_of_week=i_row - 2,
                                lesson_number=i_col - 2
                            )
                        else:
                            raise TypeError(f'В расписание добавлен не зарегистрированный пользователь - <b>{cell}</b>. '
                                            'Проверьте файл')
        except Exception as e:
            bot.send_message(message.chat.id, f"❌ Ошибка при обработке файла: {str(e)}", reply_markup=go_to_menu(), parse_mode='HTML')
            bot.set_state(message.from_user.id, reg_states_admin.admin_menu, message.chat.id)
        else:
            output_str = f'Данные успешно добавлены✅\n\nРасписание загруженной (<b>{monday_data.strftime("%d.%m.%Y")}</b>) недели:\n'
            lessons_list = Lesson.select().where(Lesson.weekly_schedule == curr_week)
            lesson_dict = {}
            for i_less in lessons_list:
                if i_less.days_dict.get(i_less.day_of_week) in lesson_dict:
                    lesson_dict[i_less.days_dict.get(i_less.day_of_week)].append([
                        i_less.lessons_dict.get(i_less.lesson_number),
                        i_less.client.clients_name,
                        i_less.client.clients_sirname])
                else:
                    lesson_dict[i_less.days_dict.get(i_less.day_of_week)] = []
                    lesson_dict[i_less.days_dict.get(i_less.day_of_week)].append([
                        i_less.lessons_dict.get(i_less.lesson_number),
                        i_less.client.clients_name,
                        i_less.client.clients_sirname])
            for i_day, lessons in lesson_dict.items():
                output_str += '\n' + f'<b>{i_day}</b>' + '\n'
                for i_less in lessons:
                    output_str += f'{i_less[0]} - {i_less[1]} {i_less[2]}\n'
            bot.send_message(message.chat.id, output_str, reply_markup=go_to_menu(), parse_mode='HTML')
            bot.set_state(message.from_user.id, reg_states_admin.admin_menu, message.chat.id)

    @bot.message_handler(state=reg_states_admin.in_schedule,
                         func=lambda message: 'Посмотреть расписание' in message.text)
    def show_schedule(message: Message):
        """
        Обработчик состояния reg_states_admin.in_schedule при нажатии админа на кнопку "Посмотреть расписание" в меню расписания.
        Предлагает админу прислать дату понедельника недели, которую тот хочет посмотреть.
        Меняет состояние пользователя на reg_states_admin.show_schedule
        :param message: сообщения от пользователя
        :return: None
        """

        bot.send_message(message.chat.id,
                         'Пришлите дату понедельника недели, расписание которой хотите посмотреть\n'
                         'Формат даты: <b>DD.MM.YYYY</b>',
                         reply_markup=go_to_menu(), parse_mode='HTML')
        bot.set_state(message.from_user.id, reg_states_admin.show_schedule, message.chat.id)

    @bot.message_handler(state=reg_states_admin.show_schedule)
    def show_schedule(message: Message):
        """
        Обработчик состояния пользователя reg_states_admin.show_schedule после отправки пользователем даты понедельника недели для показа.
        Проверяет корректность формата даты.
        В случае получения корректной даты и наличия недели по ней в БД выводит на экран расписание запрашиваемой недели.
        Меняет состояние пользователя на reg_states_admin.admin_menu
        :param message: сообщения от пользователя
        :return: None
        """

        check_format_list = message.text.split('.')
        try:
            if len(check_format_list) == 3 and 0 < int(check_format_list[0]) <= 31 and 0 < int(check_format_list[1]) <= 12 and int(check_format_list[2]) > 0:
                formated_data = datetime.strptime(message.text, "%d.%m.%Y").date()
                curr_week = Week.get_or_none(Week.monday_date == formated_data)
                lesson_list = list(curr_week.lessons)
                sorted_lesson_list = sorted(lesson_list, key=lambda i_lesson: (i_lesson.day_of_week, i_lesson.lesson_number))
                output_str = f'🕐Расписание текущей (<b>{curr_week}</b>) недели:\n'
                lesson_dict = {}
                for i_less in sorted_lesson_list:
                    if i_less.days_dict.get(i_less.day_of_week) in lesson_dict:
                        lesson_dict[i_less.days_dict.get(i_less.day_of_week)].append([
                            i_less.lessons_dict.get(i_less.lesson_number),
                            i_less.client.clients_name,
                            i_less.client.clients_sirname])
                    else:
                        lesson_dict[i_less.days_dict.get(i_less.day_of_week)] = []
                        lesson_dict[i_less.days_dict.get(i_less.day_of_week)].append([
                            i_less.lessons_dict.get(i_less.lesson_number),
                            i_less.client.clients_name,
                            i_less.client.clients_sirname])
                for i_day, lessons in lesson_dict.items():
                    output_str += '\n' + f'<b>{i_day}</b>' + '\n'
                    for i_less in lessons:
                        output_str += f'{i_less[0]} - {i_less[1]} {i_less[2]}\n'
                bot.send_message(message.chat.id,
                                 output_str, reply_markup=go_to_menu(), parse_mode='HTML')
                bot.set_state(message.from_user.id, reg_states_admin.admin_menu, message.chat.id)
            else:
                raise TypeError('Введите данные формата <b>DD.MM.YYYY</b>')
        except Exception as e:
            bot.send_message(message.chat.id,f'{e} - Введите данные формата <b>DD.MM.YYYY</b>', parse_mode='HTML')
# ========КОНЕЦ БЛОКА ПОКАЗ РАСПИСАНИЯ===========